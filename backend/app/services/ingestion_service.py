"""File ingestion: parse CSV/XLSX/JSON into metric points with auto schema detection.

The parser is intentionally defensive: it sniffs the timestamp and value
columns by name, coerces currency-ish numeric strings ("$1,234.50") and
percentages, parses a wide range of timestamp formats, and drops rows that
cannot be coerced. The result is a dataset-ready list of points plus a schema
report the UI can surface back to the user.
"""

from __future__ import annotations

import csv
import io
import json
import re
from datetime import datetime, timezone
from statistics import median
from typing import Any

from openpyxl import load_workbook

from app.utils.time import to_utc

_TS_KEYWORDS = ("date", "time", "ts", "timestamp", "datetime", "period", "day", "month")
_VALUE_KEYWORDS = (
    "value",
    "amount",
    "revenue",
    "sales",
    "quantity",
    "count",
    "metric",
    "measure",
    "total",
    "price",
    "volume",
)

_MAX_POINTS = 50_000
_MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB

_NUMBER_CLEANER = re.compile(r"[^0-9eE+\-.]")


class IngestionError(Exception):
    pass


def _clean_number(raw: Any) -> float | None:
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text in ("", "-", "n/a", "null", "None"):
        return None
    cleaned = _NUMBER_CLEANER.sub("", text.replace(",", "").replace("$", "").replace("%", ""))
    try:
        return float(cleaned)
    except ValueError:
        return None


def _parse_ts(raw: Any) -> datetime | None:
    if isinstance(raw, datetime):
        return raw
    text = str(raw).strip()
    if not text:
        return None
    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%d-%b-%Y",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%fZ",
    ]
    for fmt in candidates:
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _pick_columns(headers: list[str]) -> tuple[str | None, str | None]:
    lowered = [h.strip().lower() for h in headers]
    ts_idx = None
    value_idx = None
    for i, name in enumerate(lowered):
        if ts_idx is None and any(kw in name for kw in _TS_KEYWORDS):
            ts_idx = i
        if value_idx is None and any(kw in name for kw in _VALUE_KEYWORDS):
            value_idx = i
    if ts_idx is None:
        ts_idx = 0
    if value_idx is None:
        value_idx = 1 if len(headers) > 1 else 0
    return headers[ts_idx], headers[value_idx]


def _infer_granularity(timestamps: list[datetime]) -> str:
    if len(timestamps) < 2:
        return "day"
    ordered = sorted(timestamps)
    deltas = [
        (ordered[i + 1] - ordered[i]).total_seconds() for i in range(len(ordered) - 1)
    ]
    typical = median(deltas) if deltas else 0.0
    if typical < 3600 * 24:
        return "hour"
    if typical <= 3600 * 24 * 2:
        return "day"
    if typical <= 3600 * 24 * 10:
        return "week"
    if typical <= 3600 * 24 * 40:
        return "month"
    return "month"


def parse_csv_bytes(filename: str, raw: bytes) -> tuple[list[dict], list[str]]:
    if len(raw) > _MAX_FILE_BYTES:
        raise IngestionError("file exceeds the 10 MB limit")
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        raise IngestionError("file is empty")
    headers = [h.strip() for h in rows[0]]
    if len(rows) < 2:
        raise IngestionError("file has no data rows")
    records = [
        {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
        for r in rows[1:]
    ]
    return records, headers


def parse_xlsx_bytes(filename: str, raw: bytes) -> tuple[list[dict], list[str]]:
    if len(raw) > _MAX_FILE_BYTES:
        raise IngestionError("file exceeds the 10 MB limit")
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise IngestionError("file is empty")
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    records = [
        {headers[i]: (r[i] if i < len(r) else "") for i in range(len(headers))}
        for r in rows[1:]
    ]
    return records, headers


def parse_json_bytes(filename: str, raw: bytes) -> tuple[list[dict], list[str]]:
    try:
        data = json.loads(raw.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise IngestionError(f"invalid JSON: {exc}") from exc
    if isinstance(data, dict) and not any(isinstance(v, dict) for v in data.values()):
        records = [{"timestamp": k, "value": v} for k, v in data.items()]
        return records, ["timestamp", "value"]
    if isinstance(data, list) and all(isinstance(r, dict) for r in data):
        headers = list(data[0].keys())
        return data, headers
    raise IngestionError("JSON must be a list of objects or a {timestamp: value} map")


def parse_upload(filename: str, raw: bytes) -> dict[str, Any]:
    """Parse an uploaded file into points + schema report."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        records, headers = parse_csv_bytes(filename, raw)
    elif ext in ("xlsx", "xls"):
        records, headers = parse_xlsx_bytes(filename, raw)
    elif ext == "json":
        records, headers = parse_json_bytes(filename, raw)
    else:
        raise IngestionError("unsupported file type; use .csv, .xlsx, or .json")

    ts_col, value_col = _pick_columns(headers)
    points: list[dict[str, Any]] = []
    dropped = 0
    seen: set[datetime] = set()
    for row in records:
        ts = _parse_ts(row.get(ts_col, ""))
        value = _clean_number(row.get(value_col, ""))
        if ts is None or value is None:
            dropped += 1
            continue
        ts = to_utc(ts)
        if ts in seen:
            continue
        seen.add(ts)
        meta = {k: v for k, v in row.items() if k not in (ts_col, value_col) and v not in ("", None)}
        points.append({"timestamp": ts, "value": value, "meta": meta if meta else None})
        if len(points) >= _MAX_POINTS:
            break

    if not points:
        raise IngestionError(
            "no usable rows: could not find a timestamp column and a numeric value column"
        )

    granularity = _infer_granularity([p["timestamp"] for p in points])
    return {
        "filename": filename,
        "columns": headers,
        "timestamp_column": ts_col,
        "value_column": value_col,
        "granularity": granularity,
        "point_count": len(points),
        "dropped": dropped,
        "points": points,
    }


def build_payload_from_records(points: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{"timestamp": p["timestamp"], "value": p["value"], "meta": p.get("meta")} for p in points]
